import copy
from pathlib import Path

import openpyxl
import yaml
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from c3hm.commands.rubric import export_rubric_data
from c3hm.data.student import Student, find_student_by_name, read_omnivox_students_file


class FeedBackStudent:
    """
    Classe représentant un étudiant pour la génération de rétroaction.
    Contient les informations présentes dans le fichier de rétroaction.
    """
    def __init__(self, name: str, matricule: str, grade: float, comment: str):
        self.name = name
        self.matricule = matricule
        self.grade = grade
        self.comment = comment

def generate_feedback(gradebook_path: Path, output_dir: Path, students_file: Path | None):
    """
    Génère un document Excel de rétroaction pour les étudiants à partir d’une fichier de correction
    et un résumé des notes en format Excel.
    """

    # Génère le fichier Excel pour charger les notes dans Omnivox
    students = read_omnivox_students_file(students_file) if students_file else None
    students = process_yaml_files(gradebook_path, output_dir, students)
    generate_xl_for_omnivox(students, output_dir)


def process_yaml_files(
    gradebook_path: Path,
    output_dir: Path | str,
    student_list: list[Student] | None
) -> list[FeedBackStudent]:
    """
    Pour chaque fichier de correction dans le répertoire, génère un fichier PDF
    """
    output_dir = Path(output_dir)
    if not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)

    yaml_files = list(gradebook_path.glob("*.yaml"))
    all_students: list[FeedBackStudent] = []
    for yaml_file in yaml_files:
        try:
            with open(yaml_file, encoding="utf-8") as f:
                data = yaml.safe_load(f)

            students: list[tuple[str, str]] = []
            if "étudiant" in data:
                students.append(extract_student(data["étudiant"], student_list))
            elif "étudiants" in data:
                for s in data["étudiants"]:
                    if s["nom"] is not None or s.get("matricule") is not None:
                        students.append(extract_student(s, student_list))
            else:
                raise ValueError("Le fichier YAML doit contenir une section 'étudiant' ou 'étudiants'.")

            if not students:
                print(f"Aucun étudiant trouvé dans le fichier '{yaml_file}', aucun fichier de rétroaction généré.")
                continue

            for (name, matricule) in students:
                destination = output_dir / f"{name} {matricule}.pdf"
                data_student = copy.deepcopy(data)
                data_student["nom"] = name
                data_student["matricule"] = matricule
                grade = 0.0
                for node in data_student["critères"]:
                    if "section" in node:
                        continue
                    if "pourcentage" not in node:
                        raise ValueError("Chaque critère doit contenir un pourcentage.")
                    node["pourcentage"] = parse_percent(node["pourcentage"])
                    node["note"] = round(node["pourcentage"] * node["points"], 1)
                    grade += node["note"]
                bonus_malus = data_student.get("bonus malus", {})
                if bonus_malus.get("points") is not None:
                    grade += bonus_malus["points"]
                data_student["note"] = round(grade, 0)
                export_rubric_data(data_student, destination)
                student = FeedBackStudent(
                    name=name,
                    matricule=matricule,
                    grade=grade,
                    comment="")
                all_students.append(student)
        except Exception as e:
            raise RuntimeError(f"Erreur lors de la génération des fichiers de rétroaction pour le fichier '{yaml_file}'") from e

    return all_students

def extract_student(s: dict, student_list: list[Student] | None) -> tuple[str, str]:
    if "matricule" not in s:
        if not student_list:
            raise ValueError("Le fichier d'étudiants doit être fourni pour faire la correspondance par nom.")
        s1 = find_student_by_name(s["nom"], student_list)
        return s1.full_name(), s1.omnivox_id
    else:
        return s["nom"], s["matricule"]

def generate_xl_for_omnivox(
    students: list[FeedBackStudent],
    output_dir: Path | str
) -> None:
    """
    Génère un fichier Excel pour charger les notes dans Omnivox.
    """
    output_dir = Path(output_dir)
    if not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)
    omnivox_path = output_dir / "notes_omnivox.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    populate_omnivox_sheet(students, ws)

    # Sauvegarde le fichier Excel
    wb.save(omnivox_path)

def populate_omnivox_sheet(students: list[FeedBackStudent], omnivox_worksheet: Worksheet) -> None:
    omnivox_worksheet.title = "Notes pour Omnivox"
    omnivox_worksheet.sheet_view.showGridLines = False  # Disable gridlines

    # En-têtes
    omnivox_worksheet.append(["Code omnivox", "Note", "Commentaire", "Nom"])

    # Trouves tous les fichiers excel
    for student in students:
        omnivox_worksheet.append([student.matricule, student.grade, student.comment, student.name])

    # Format
    _insert_table(omnivox_worksheet, "NotesOmnivox", "A1:D" + str(omnivox_worksheet.max_row))
    omnivox_worksheet.column_dimensions["A"].width = 20
    omnivox_worksheet.column_dimensions["B"].width = 10
    omnivox_worksheet.column_dimensions["C"].width = 70
    omnivox_worksheet.column_dimensions["D"].width = 40

def parse_percent(note: str | float | int | None) -> float:
    if note is None:
        raise ValueError("La note ne peut pas être None")
    if isinstance(note, float | int):
        grade = float(note)
    elif isinstance(note, str):
        note = note.strip().lower()
        if note in ("tb", "très bien", "tres bien"):
            grade =  1.0
        elif note in ("b", "bien"):
            grade =  0.80
        elif note in ("p", "passable"):
            grade =  0.6
        elif note in ("a", "à améliorer", "a ameliorer"):
            grade =  0.30
        elif note in ("i", "insuffisant"):
            grade =  0.0
        else:
            grade =  float(note)
    else:
        raise TypeError(f"Type de note inattendu: {type(note)}")
    if not (0.0 <= grade <= 1.0):
        raise ValueError(f"La note doit être entre 0 et 1. Valeur reçue: {note}")
    return grade

def _insert_table(ws: Worksheet, display_name: str, ref: str) -> None:
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)
