from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import absolute_coordinate, get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_GRID_SIZE = 4  # Nombre par défaut de critères et d'indicateurs

PERFECT_GREEN    = "C8FFC8"  # RGB(200, 255, 200)
VERY_GOOD_GREEN  = "F0FFB0"  # RGB(240, 255, 176)
HALF_WAY_YELLOW  = "FFF8C2"  # RGB(255, 248, 194)
MINIMAL_RED      = "FFE4C8"  # RGB(255, 228, 200)
BAD_RED          = "FFC8C8"  # RGB(255, 200, 200)

GRADE_COLORS = {
    2: [PERFECT_GREEN, BAD_RED],
    3: [PERFECT_GREEN, HALF_WAY_YELLOW, BAD_RED],
    4: [PERFECT_GREEN, VERY_GOOD_GREEN, HALF_WAY_YELLOW, BAD_RED],
    5: [PERFECT_GREEN, VERY_GOOD_GREEN, HALF_WAY_YELLOW, MINIMAL_RED, BAD_RED],
}

GRADE_LEVELS = {
    2: ["Réussi", "Insuffisant"],
    3: ["Très bien", "Acceptable", "Insuffisant"],
    4: ["Très bien", "Bien", "Passable", "Insuffisant"],
    5: ["Très bien", "Bien", "Passable", "À améliorer", "Insuffisant"],
}

GRADE_PERCENTAGES = {
    2: [1.0, 0],
    3: [1.0, 0.6, 0],
    4: [1.0, 0.8, 0.6, 0],
    5: [1.0, 0.8, 0.6, 0.3, 0],
}

TEAM_WS_NAME = "Équipe"
STUDENT_WS_PREFIX = "Étudiant "

def distribute_points(total_points: int, num_items: int) -> list[int]:
    """
    Distribue un nombre total de points équitablement entre un nombre d'items.
    Les valeurs ne diffèrent pas de plus de 1.
    """
    base_value = total_points // num_items
    remainder = total_points % num_items

    # Les 'remainder' premiers items reçoivent base_value + 1, les autres base_value
    points = [base_value + 1 if i < remainder else base_value for i in range(num_items)]

    return points


def export_template(output_path: Path,
                    nb_levels: int = 5,
                    team_size: int = 1,
                    criteria_indicators: list[int] | None = None) -> None:
    """
    Génère une grille d'évaluation sous format Excel.
    """
    # Valider les paramètres
    if criteria_indicators is None:
        criteria_indicators = [DEFAULT_GRID_SIZE] * DEFAULT_GRID_SIZE
    check_template_params(nb_levels, criteria_indicators, team_size)

    # Infèrer des valeurs utiles
    nb_criteria = len(criteria_indicators)
    nb_indicators = sum(criteria_indicators)
    indicator_points = distribute_points(100, nb_indicators)

    grade_col = 4 + nb_levels
    grade_letter = get_column_letter(grade_col)
    comment_col = grade_col + 1
    criteria_row = 10

    # Create new workbook
    wb = create_workbook(team_size)

    for ws in wb.worksheets:
        set_columns_width(ws, nb_levels)
        create_front_matter(wb, ws, grade_letter, criteria_indicators, criteria_row)

        # Create grid
        create_grid(wb, ws, nb_levels, criteria_indicators,
                    indicator_points, nb_criteria,
                    grade_col, comment_col,
                    criteria_row)

    # Save workbook
    wb.save(output_path)

def create_workbook(team_size: int) -> Workbook:
    wb = Workbook()
    if team_size == 1:
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Grille"
    else:
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = TEAM_WS_NAME
        for i in range(team_size):
            ws = wb.create_sheet(title=f"{STUDENT_WS_PREFIX}{i + 1}")
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    return wb

def get_current_semester() -> str:
    """
    Retourne le semestre actuel
    """
    today = date.today()
    year = today.year
    if today.month <= 5:
        return f"Hiver {year}"
    elif today.month <= 7:
        return f"Été {year}"
    else:
        return f"Automne {year}"

def set_header_style(cell: Cell):
    cell.style = "Headline 4"
    cell.alignment = Alignment(horizontal="right")

def all_indicators_range(col_letter: str, criteria_indicators: list[int],
                         criteria_row: int,
                         include_penalty: bool = True) -> str:
    """
    Retourne la plage Excel couvrant tous les indicateurs, et
    éventuellement la cellule de pénalité.
    """
    r = []
    for criterion_idx, nb_indicators in enumerate(criteria_indicators):
        row_start = criteria_row + 1 + sum(criteria_indicators[:criterion_idx]) + criterion_idx
        row_end = row_start + nb_indicators - 1
        r.append(f"{col_letter}{row_start}:{col_letter}{row_end}")
    total_indicators = sum(criteria_indicators)
    if include_penalty:
        nb_criteria = len(criteria_indicators)
        penalty_row = criteria_row + total_indicators + nb_criteria + 2
        r.append(f"{col_letter}{penalty_row}")
    return ",".join(r)

def check_template_params(nb_levels: int,
                          criteria_indicators: list[int] | None,
                          team_size: int) -> None:
    """
    Valide les paramètres pour la génération de la grille d'évaluation.
    """
    if nb_levels < 2 or nb_levels > 5:
        raise ValueError("Le nombre de niveaux doit être entre 2 et 5.")

    if not criteria_indicators:
        raise ValueError("Il faut au moins un critère.")

    if any(n < 1 for n in criteria_indicators):
        raise ValueError("Chaque critère doit avoir au moins un indicateur.")

    if team_size < 1:
        raise ValueError("La taille de l'équipe doit être au moins 1.")

def create_defined_name(wb: Workbook, ws: Worksheet, name: str, cell_coord: str) -> None:
    """
    Crée un nom défini au niveau de la feuille Excel.
    """
    ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell_coord)}"
    local_sheet_id = wb.worksheets.index(ws) # type: ignore
    defn = DefinedName(name, attr_text=ref, localSheetId=local_sheet_id)
    ws.defined_names[name] = defn

def create_front_matter(
    wb: Workbook,
    ws: Worksheet,
    grade_letter: str,
    criteria_indicators: list[int],
    criteria_row: int,
) -> None:
    """
    Crée la partie avant la grille dans la feuille Excel (nom, matricule, note, commentaire).
    """
    team_ws = wb[TEAM_WS_NAME] if ws.title.startswith(STUDENT_WS_PREFIX) else None

    # Nom et matricule uniquement si ce n'est pas la feuille équipe
    if ws.title != TEAM_WS_NAME:
        ws.cell(row=2, column=2, value="Matricule")
        set_header_style(ws.cell(row=2, column=2)) # type: ignore
        create_defined_name(wb, ws, "cthm_matricule", "C2")

        ws.cell(row=2, column=4, value="Nom")
        set_header_style(ws.cell(row=2, column=4)) # type: ignore
        create_defined_name(wb, ws, "cthm_nom", "E2")

    # Note globale
    ws.cell(row=3, column=2, value="Note")
    set_header_style(ws.cell(row=3, column=2)) # type: ignore
    create_defined_name(wb, ws, "cthm_note", "C3")
    ws.cell(row=3, column=3,
            value=f'=_xlfn.CONCAT(SUM({all_indicators_range(grade_letter,
                                                            criteria_indicators,
                                                            criteria_row=criteria_row)})," points")')

    # Commentaire global
    ws.cell(row=3, column=4, value="Commentaire")
    set_header_style(ws.cell(row=3, column=4)) # type: ignore
    if team_ws:
        ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate("E3")}'
        ws.cell(row=3, column=5,
                value=f'={if_blank_null_str(ref)}')
    create_defined_name(wb, ws, "cthm_commentaire", "E3")

    # Session
    ws.cell(row=5, column=2, value="Session")
    set_header_style(ws.cell(row=5, column=2)) # type: ignore
    if team_ws:
        ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate("C5")}'
        ws.cell(row=5, column=3,
                value=f'={if_blank_null_str(ref)}')
    else:
        ws.cell(row=5, column=3, value=f"{get_current_semester()}")

    # Cours
    ws.cell(row=5, column=4, value="Cours")
    set_header_style(ws.cell(row=5, column=4)) # type: ignore
    if team_ws:
        ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate("E5")}'
        ws.cell(row=5, column=5,
                value=f'={if_blank_null_str(ref)}')

    # Évaluation
    ws.cell(row=5, column=6, value="Évaluation")
    set_header_style(ws.cell(row=5, column=6)) # type: ignore
    if team_ws:
        ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate("G5")}'
        ws.cell(row=5, column=7,
                value=f'={if_blank_null_str(ref)}')

def if_blank_null_str(cell_coord: str) -> str:
    return f'IF(ISBLANK({cell_coord}),"",{cell_coord})'

def set_columns_width(ws: Worksheet, nb_levels: int) -> None:
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 13
    for i in range(nb_levels):
        col_letter = chr(ord("D") + i)
        ws.column_dimensions[col_letter].width = 16
    ws.column_dimensions[chr(ord("D") + nb_levels)].width = 12  # Grade column
    ws.column_dimensions[chr(ord("D") + nb_levels + 1)].width = 70  # Comment column

def create_grid(
    wb: Workbook,
    ws: Worksheet,
    nb_levels: int,
    criteria_indicators: list[int],
    indicator_points: list[int],
    nb_criteria: int,
    grade_col: int,
    comment_col: int,
    criteria_row: int,
) -> None:
    """
    Crée la grille d'évaluation dans la feuille Excel.
    """
    team_ws = wb[TEAM_WS_NAME] if ws.title.startswith(STUDENT_WS_PREFIX) else None

    grade_letter = get_column_letter(grade_col)

    ws.cell(row=8, column=3,
            value=f'=_xlfn.CONCAT("Total sur ",SUM({all_indicators_range("C", criteria_indicators,
                                                                         criteria_row=criteria_row,
                                                                         include_penalty=False)})," points")')
    ws.cell(row=8, column=3).style = "Explanatory Text"
    total_str = "La note pour chaque critère est soit "
    for level_idx in range(nb_levels):
        perc = int(GRADE_PERCENTAGES[nb_levels][level_idx] * 100)
        if level_idx == nb_levels - 1:
            total_str += f"{perc}%"
        else:
            total_str += f"{perc}%, "
    total_str += " des points. Le professeur peut ajuster exceptionnellement."
    ws.cell(row=9, column=3, value=total_str)
    ws.cell(row=9, column=3).style = "Explanatory Text"
    for criterion_idx in range(nb_criteria):
        # Déterminer la ligne du critère
        nb_indicators = criteria_indicators[criterion_idx]
        criterion_row = criteria_row + sum(criteria_indicators[:criterion_idx]) + criterion_idx
        if team_ws:
            ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate("B" + str(criterion_row))}'
            ws.cell(row=criterion_row, column=2,
                    value=f'={ref}')
        else:
            ws.cell(row=criterion_row, column=2,
                    value=f"Critère {criterion_idx + 1}")
        ws.cell(row=criterion_row, column=2).style = "Headline 1"

        pts_range = f"C{criterion_row + 1}:C{criterion_row + nb_indicators}"
        ws.cell(row=criterion_row, column=3, value=f'=_xlfn.CONCAT(SUM({pts_range})," points")')
        ws.cell(row=criterion_row, column=3).style = "Explanatory Text"

        for level_idx in range(nb_levels):
            perc = f"{int(GRADE_PERCENTAGES[nb_levels][level_idx] * 100)}"
            ws.cell(row=criterion_row,
                    column=4 + level_idx,
                    value=f"{GRADE_LEVELS[nb_levels][level_idx]} ({perc}%)")
            color = GRADE_COLORS[nb_levels][level_idx]
            cell = ws.cell(row=criterion_row, column=4 + level_idx)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        pts_range = f"{grade_letter}{criterion_row + 1}:{grade_letter}{criterion_row + nb_indicators}"
        ws.cell(row=criterion_row, column=grade_col,
                value=f'=_xlfn.CONCAT("Note : ",SUM({pts_range}))')

        ws.cell(row=criterion_row, column=comment_col, value="Commentaire")

        # Create indicators
        for indicator_idx in range(nb_indicators):
            indicator_row = criterion_row + indicator_idx + 1
            # Calculer l'index global de l'indicateur
            global_indicator_idx = sum(criteria_indicators[:criterion_idx]) + indicator_idx

            if team_ws:
                ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate("B" + str(indicator_row))}'
                ws.cell(row=indicator_row, column=2,
                        value=f'={ref}')
                ref_points = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate("C" + str(indicator_row))}'
                ws.cell(row=indicator_row, column=3,
                        value=f'={ref_points}')
            else:
                ws.cell(row=indicator_row, column=2,
                        value=f"Indicateur {criterion_idx + 1}.{indicator_idx + 1}")
                ws.cell(row=indicator_row, column=3, value=indicator_points[global_indicator_idx])
            ws.cell(row=indicator_row, column=3).style = "Explanatory Text"

            # Centrer les cellules de niveau de performance
            for level_idx in range(nb_levels):
                cell = ws.cell(row=indicator_row, column=4 + level_idx)
                cell.alignment = Alignment(horizontal="center")

            # Insert grade formula
            xs = []
            for level_idx in range(nb_levels):
                cell_letter = chr(ord("D") + level_idx)
                cell_coord = f"{cell_letter}{indicator_row}"
                default_percent = GRADE_PERCENTAGES[nb_levels][level_idx]
                grade_or_percent = (f'IF(LEFT(_xlfn.CELL("format", {cell_coord}),1)="P",'
                                    f'C{indicator_row}*{cell_coord},{cell_coord})')
                xs.append(f"IF(ISTEXT({cell_coord}),C{indicator_row}*{default_percent},{grade_or_percent})")
            counta_range = f"D{indicator_row}:{chr(ord('D') + nb_levels - 1)}{indicator_row}"
            if team_ws:
                team_ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate(grade_letter + str(indicator_row))}'
                grade = f"IF(COUNTA({counta_range})=0,{team_ref},{'+'.join(xs)})"
            else:
                grade = "+".join(xs)
            formula = f"=IF(COUNTA({counta_range})>1,NA(),{grade})"
            ws.cell(row=indicator_row, column=grade_col, value = formula)

            # Insert comment formula if needed
            if team_ws:
                ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate(get_column_letter(comment_col) + str(indicator_row))}'
                ws.cell(row=indicator_row, column=comment_col,
                        value=f'={if_blank_null_str(ref)}')

    # Pénalités pour retard, français, fautes significatives
    nb_indicators = sum(criteria_indicators)
    penalty_row = criteria_row + nb_indicators + nb_criteria + 1
    ws.cell(row=penalty_row, column=3 + nb_levels + 1, value="Points")
    if team_ws:
        ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate(get_column_letter(3 + nb_levels + 1) + str(penalty_row + 1))}'
        ws.cell(row=penalty_row + 1, column=3 + nb_levels + 1,
                value=f'={ref}')
    else:
        ws.cell(row=penalty_row + 1, column=3 + nb_levels + 1, value=0)

    ws.cell(row=penalty_row, column=3 + nb_levels + 2, value="Commentaire")
    if team_ws:
        ref = f'{quote_sheetname(team_ws.title)}!{absolute_coordinate(get_column_letter(3 + nb_levels + 2) + str(penalty_row + 1))}'
        ws.cell(row=penalty_row + 1, column=3 + nb_levels + 2,
                value=f'={if_blank_null_str(ref)}')

    ws.cell(row=penalty_row + 1, column=2, value="Bonus / Malus")
    ws.cell(row=penalty_row + 1, column=2).style = "Headline 1"

    ws.cell(row=penalty_row + 2, column=2,
            value="En plus de la grille ci-dessus, il est possible "
                  "que des points soient retirés pour :")
    ws.cell(row=penalty_row + 3, column=2, value="- un retard")
    ws.cell(row=penalty_row + 4, column=2, value="- des fautes de français")
    ws.cell(row=penalty_row + 5, column=2,
            value="- une erreur significative (non respect "
                  "des conventions d'usage, absence de commentaires lorsque nécessaire, "
                  "code spaghetti, code qui plante ou ne démarre pas, etc.)")
    for i in range(4):
        ws.cell(row=penalty_row + 2 + i, column=2).style = "Explanatory Text"
