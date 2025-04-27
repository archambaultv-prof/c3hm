from pathlib import Path

import openpyxl as pyxl
import openpyxl.utils as pyxl_utils
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from c3hm.core.generate.generate_word import scale_color_schemes
from c3hm.core.rubric import Rubric, Student


def generate_excel_from_rubric(rubric: Rubric, output_path: Path | str) -> None:
    """
    Génère un document Excel servant à la correction à partir d’une Rubric
    """
    wb = pyxl.Workbook()
    for sheet in wb.worksheets:
        wb.remove(sheet)

    ws = wb.create_sheet(title="c3hm")
    add_c3hm_sheet(ws, rubric)

    for student in rubric.students:
        # Crée une feuille pour chaque étudiant
        ws = wb.create_sheet(title=student.id)
        add_student_sheet(ws, rubric, student)

    wb.save(output_path)

def add_c3hm_sheet(ws: Worksheet, rubric: Rubric) -> None:
    ws.sheet_view.showGridLines = False

    # Tailles de colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30

    # Info étudiant
    ws.append(["Code Omnivox", "Prénom", "Nom de famille", "Id"])
    for i in range(1, 5):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.style = "Headline 1"
    cell = ws.cell(row=ws.max_row, column=1)
    dn = DefinedName(
        name="cthm_code_omnivox",
        attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell.coordinate)}",
    )
    ws.defined_names.add(dn)
    for student in rubric.students:
        ws.append([student.omnivox_code, student.first_name, student.last_name, student.id])

    # Paramètres de la grille
    key_col = 6
    key_row = 1
    ws.cell(row=1, column=key_col).value = "POUR RÉFÉRENCE, NE PAS CHANGER CES PARAMÈTRES"
    ws.cell(row=1, column=key_col).style = "Warning Text"

    key_row = 2
    ws.cell(row=key_row, column=key_col).value = "Clé"
    ws.cell(row=key_row, column=key_col).style = "Headline 1"
    ws.cell(row=key_row, column=key_col + 1).value = "Valeur"
    ws.cell(row=key_row, column=key_col + 1).style = "Headline 1"
    cell = ws.cell(row=key_row, column=key_col + 1)
    dn = DefinedName(
        name="cthm_config_key",
        attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell.coordinate)}",
    )
    ws.defined_names.add(dn)

    ws.cell(row=key_row + 1, column=key_col).value = "Cours"
    ws.cell(row=key_row + 1, column=key_col + 1).value = rubric.course

    ws.cell(row=key_row + 2, column=key_col).value = "Évaluation"
    ws.cell(row=key_row + 2, column=key_col + 1).value = rubric.evaluation

    ws.cell(row=key_row + 3, column=key_col).value = "Total"
    ws.cell(row=key_row + 3, column=key_col + 1).value = rubric.grid.total_score

    ws.cell(row=key_row + 4, column=key_col).value = "Précision poids"
    ws.cell(row=key_row + 4, column=key_col + 1).value = rubric.grid.pts_precision

    ws.cell(row=key_row + 5, column=key_col).value = "Précision seuils"
    ws.cell(row=key_row + 5, column=key_col + 1).value = rubric.grid.thresholds_precision

    level_row = key_row + 6
    for i, level in enumerate(rubric.grid.scale):
        ws.cell(row=level_row + i, column=key_col).value = f"Niveau {i+1}"
        ws.cell(row=level_row + i, column=key_col + 1).value = level

    threshold_row = level_row + len(rubric.grid.scale)
    for i, threshold in enumerate(rubric.grid.thresholds):
        ws.cell(row=threshold_row + i, column=key_col).value = f"Seuil {i+1}"
        ws.cell(row=threshold_row + i, column=key_col + 1).value = threshold

    next_row = threshold_row + len(rubric.grid.thresholds)
    for criterion in rubric.grid.criteria:
        ws.cell(row=next_row, column=key_col).value = "Critère"
        ws.cell(row=next_row, column=key_col).style = "Headline 3"
        ws.cell(row=next_row, column=key_col + 1).value = criterion.name
        ws.cell(row=next_row, column=key_col + 1).style = "Headline 3"

        ws.cell(row=next_row + 1, column=key_col).value = "Critère poids"
        ws.cell(row=next_row + 1, column=key_col + 1).value = criterion.points

        next_row += 2
        for indicator in criterion.indicators:
            ws.cell(row=next_row, column=key_col).value = "Indicateur"
            ws.cell(row=next_row, column=key_col).style = "Headline 4"
            ws.cell(row=next_row, column=key_col + 1).value = indicator.name
            ws.cell(row=next_row, column=key_col + 1).style = "Headline 4"

            next_row += 1
            for idx, descriptor in enumerate(indicator.descriptors):
                ws.cell(row=next_row, column=key_col).value = f"Descripteur {idx+1}"
                ws.cell(row=next_row, column=key_col).style = "Explanatory Text"
                ws.cell(row=next_row, column=key_col + 1).value = descriptor
                ws.cell(row=next_row, column=key_col + 1).style = "Explanatory Text"
                next_row += 1

def add_student_sheet(ws: Worksheet, rubric: Rubric, student: Student) -> None:
    ws.sheet_view.showGridLines = False

    # Tailles de colonnes
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    for i in range(3, 11):
        ws.column_dimensions[pyxl_utils.get_column_letter(i)].width = 15
    ws.column_dimensions[pyxl_utils.get_column_letter(11)].width = 60

    # Titre
    ws.append([rubric.title()])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.style = "Title"
    ws.append([])

    # Info étudiant
    ws.append(["Étudiant", f"{student.first_name} {student.last_name}"])
    ws.append(["Code omnivox", f"{student.omnivox_code}"])
    cell = ws.cell(row=ws.max_row, column=2)
    dn = DefinedName(
        name="cthm_code_omnivox",
        attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell.coordinate)}",
    )
    ws.defined_names.add(dn)
    ws.append([])

    # Total sur X pts
    pts = "pt" if rubric.grid.total_score == 1 else "pts"
    ws.append([f"Total sur {rubric.grid.total_score} {pts}"])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.style = "Headline 2"

    total_row = ws.max_row
    total_cell = 2
    cell = ws.cell(row=ws.max_row, column=2)
    cell.style = "Calculation"
    cell.number_format = "0.0"

    # Grille
    grid = rubric.grid
    scale_colors = scale_color_schemes(len(grid.scale))

    # En-tête - seuils
    col_before_scale = 2
    ws.append([''] * col_before_scale + grid.thresholds)
    threshold_perfect_cell = ws.cell(row=ws.max_row, column=col_before_scale + 1).coordinate
    threshold_row = ws.max_row
    for i in range(col_before_scale + 1, len(grid.thresholds) + col_before_scale + 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.style = "Explanatory Text"

    # En-tête - barème
    extra_cols = ["note calculée", "note manuelle", "note en pts", "commentaires"]
    ws.append([''] * col_before_scale + grid.scale + extra_cols)
    for i in range(col_before_scale + 1, len(grid.scale) + col_before_scale + 1 + len(extra_cols)):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.style = "Headline 4"
        if scale_colors and i - col_before_scale - 1 < len(scale_colors):
            color = scale_colors[i - col_before_scale - 1]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


    pts_letter = "B"
    criterion_rows = []
    computed_col = col_before_scale + len(grid.scale) + 1
    computed_letter = pyxl_utils.get_column_letter(computed_col)
    manual_col = col_before_scale + len(grid.scale) + 2
    manual_letter = pyxl_utils.get_column_letter(manual_col)
    final_col = col_before_scale + len(grid.scale) + 3
    final_letter = pyxl_utils.get_column_letter(final_col)
    for cidx, criterion in enumerate(grid.criteria):
        # Critère
        ws.append([criterion.name, criterion.points])
        cr = ws.max_row
        criterion_rows.append(cr)
        cell = ws.cell(row=ws.max_row, column=1)
        cell.style = "Headline 3"
        dn = DefinedName(
            name=f"cthm_C{cidx+1}_label",
            attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell.coordinate)}",
        )
        ws.defined_names.add(dn)

        # note calculée
        cell = ws.cell(row=ws.max_row, column= computed_col)
        computed_range = (f"{computed_letter}{cr+1}:"
                          f"{computed_letter}{cr+len(criterion.indicators)}")
        weight_range = (f"{pts_letter}{cr+1}:"
                        f"{pts_letter}{cr+len(criterion.indicators)}")
        cell.value = (f"=sumproduct({computed_range},{weight_range}) / sum({weight_range})")
        cell.style = "Calculation"
        cell.number_format = "0.0"

        # note manuelle
        cell = ws.cell(row=ws.max_row, column= manual_col)
        cell.style = "Input"
        cell.number_format = "0.0"

        # note en pts
        cell = ws.cell(row=ws.max_row, column= col_before_scale + len(grid.scale) + 3)
        computed_or_manual = (f"IF(ISBLANK({manual_letter}{cr}),{computed_letter}{cr},"
                              f"{manual_letter}{cr})")
        cell.value = f"={computed_or_manual}/{threshold_perfect_cell}*{pts_letter}{cr}"
        cell.style = "Calculation"
        cell.number_format = "0.0"
        dn = DefinedName(
            name=f"cthm_C{cidx+1}",
            attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell.coordinate)}",
        )
        ws.defined_names.add(dn)

        # Descripteur
        for idix, indicator in enumerate(criterion.indicators):
            ws.append([indicator.name, 1])
            cell = ws.cell(row=ws.max_row, column=1)
            dn = DefinedName(
                name=f"cthm_C{cidx+1}_I{idix+1}_label",
                attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell.coordinate)}",
            )
            ws.defined_names.add(dn)
            for i in range(2):
                cell = ws.cell(row=ws.max_row, column=i+1)
                cell.style = "Explanatory Text"
            # Affichage conditionnel pour les descripteurs
            for i in range(len(indicator.descriptors)):
                color = "D3D3D3"  # Gris clair
                if scale_colors:
                    color = scale_colors[i]
                col_letter = pyxl_utils.get_column_letter(i + col_before_scale + 1)
                rule = FormulaRule(
                    formula=[f'NOT(ISBLANK({col_letter}{ws.max_row}))'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
                )
                ws.conditional_formatting.add(
                    f'{col_letter}{ws.max_row}',
                    rule
                )

            # Note pour l'indicateur.
            # Si on utilise une formule trop avancée, Excel
            # va ajouter un @ sur certains ranges, ce qui va casser la formule.
            # Donc on utilise des formules simples sans confusion avec les
            # dynamic arrays.
            r = ws.max_row
            def cell_addr(i, r=r):
                return f"{pyxl_utils.get_column_letter(i + col_before_scale + 1)}{r}"
            def thr_addr(i):
                return f"{pyxl_utils.get_column_letter(i + col_before_scale + 1)}{threshold_row}"
            one_cell = [f"IF(ISBLANK({cell_addr(i)}),0,1)"
                        for i in range(len(indicator.descriptors))]
            one_cell = f"{" + ".join(one_cell)} = 1"
            val = [f"IF(ISBLANK({cell_addr(i)}),0,"
                   f"IF(ISNUMBER({cell_addr(i)}),{cell_addr(i)},{thr_addr(i)}))"
                   for i in range(len(indicator.descriptors))]
            val = f"{" + ".join(val)}"
            cell = ws.cell(row=ws.max_row, column=computed_col)
            cell.value = (f"=IF({one_cell},{val},NA())")
            cell.style = "Output"
            cell.number_format = "0.0"
            dn = DefinedName(
                name=f"cthm_C{cidx+1}_I{idix+1}",
                attr_text=f"{quote_sheetname(ws.title)}!{absolute_coordinate(cell.coordinate)}",
            )
            ws.defined_names.add(dn)

    # Formule pour le total
    f = "=sum("
    f += ",".join([f"{final_letter}{r}" for r in criterion_rows])
    f += ")"
    ws.cell(row=total_row, column=total_cell).value = f
