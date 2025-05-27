import csv
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field


class Student(BaseModel):
    """
    Représente un étudiant avec un nom, prénom, code Omnivox et alias.
    Possibilité d'être associé à une équipe. Si l'étudiant est la référence
    pour son équipe, cela implique que la note attribuée à l'étudiant
    sera également pour les coéquipiers, sauf si une note est spécifiquement
    attribuée à un coéquipier.
    """
    model_config = ConfigDict(coerce_numbers_to_str=True)

    first_name: str = Field(..., min_length=1)
    last_name: str = Field(..., min_length=1)
    omnivox_code: str = Field(..., min_length=1)
    alias: str = Field(..., min_length=1)
    team: str | None = None
    is_team_reference: bool = False

    def has_team(self) -> bool:
        """
        Vérifie si l'étudiant est associé à une équipe.
        """
        return self.team is not None and self.team.strip() != ""

    def ws_name(self) -> str:
        """
        Retourne le nom de la feuille de calcul pour l'étudiant.
        """
        return self.alias

    def to_dict(self) -> dict:
        """
        Retourne un dictionnaire représentant l'étudiant.
        """
        return {
            "code omnivox": self.omnivox_code,
            "prénom": self.first_name,
            "nom de famille": self.last_name,
            "alias": self.alias,
            "équipe": self.team,
            "référence d'équipe": self.is_team_reference,
        }

    @classmethod
    def from_dict(cls, data: dict) -> "Student":
        """
        Crée une instance de Student à partir d'un dictionnaire.
        """
        def one_of(*args):
            for s in args:
                if s in data:
                    return data[s]
            raise KeyError(
                f"Un des champs suivants est requis: {', '.join(args)}"
            )
        if "référence d'équipe" in data:
            r = data["référence d'équipe"]
            r = False if isinstance(r, str) and r.strip() == "" or r is None else str(r)
        else:
            r = False
        return cls(
            first_name=one_of("prénom", "Prénom de l'étudiant"),
            last_name=one_of("nom de famille", "Nom de l'étudiant"),
            omnivox_code=one_of("code omnivox", "No de dossier"),
            alias=data["alias"],
            team=data.get("équipe"),
            is_team_reference=r, # type: ignore
        )

    def copy(self) -> "Student": # type: ignore
        """
        Retourne une copie de l'étudiant.
        """
        return Student(
            first_name=self.first_name,
            last_name=self.last_name,
            omnivox_code=self.omnivox_code,
            alias=self.alias,
            team=self.team,
            is_team_reference=self.is_team_reference,
        )

def read_student(path: str | Path) -> list[dict[str, str]]:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return read_student_csv(path)
    elif path.suffix.lower() in {".xlsx", ".xls"}:
        return read_student_excel(path)
    else:
        raise ValueError(
            f"Le format de fichier {path.suffix} n'est pas supporté. "
            "Utilisez un fichier CSV ou Excel."
        )

def read_student_csv(path: str | Path) -> list[dict[str, str]]:
    """
    Lit un fichier CSV contenant des informations sur les étudiants et retourne
    une liste d'instances de Student.
    """
    students = []
    with open(path, encoding="utf-8") as file:
        csv_data = csv.DictReader(file)
        for row in csv_data:
            students.append(row)
    return students

def read_student_excel(path: str | Path) -> list[dict[str, str]]:
    """
    Lit un fichier Excel contenant des informations sur les étudiants et retourne
    une liste de dictionnaires représentant chaque étudiant.
    La structure attendue est la même que pour le fichier CSV.
    """

    # We avoid pandas to keep the dependencies minimal and because
    # sometimes as Collège de Maisonneuve it is not installed.
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True)) # type: ignore
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    students = []
    for row in rows[1:]:
        student = {headers[i]: (str(cell) if cell is not None else "")
                   for i, cell in enumerate(row)}
        if is_empty_row(student):
            continue
        students.append(student)
    return students

def is_empty_row(row: dict[str, str]) -> bool:
    """
    Vérifie si une ligne de données d'étudiant est vide.
    """
    return all(value.strip() == "" for value in row.values())
