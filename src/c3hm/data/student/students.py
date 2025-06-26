import csv
from pathlib import Path
from typing import Self

from openpyxl import load_workbook
from pydantic import BaseModel, model_validator

from c3hm.data.student.student import Student


class Students(BaseModel):
    """
    Classe pour gérer les étudiants.
    """
    students: list[Student]

    @model_validator(mode="after")
    def validate_students(self) -> Self:
        """
        Valide la liste des étudiants.
        """
        # Les alias des étudiants doivent être uniques
        aliases = {student.alias for student in self.students}
        if len(aliases) != len(self.students):
            dup = {}
            for student in self.students:
                if student.alias not in dup:
                    dup[student.alias] = 0
                dup[student.alias] += 1
            duplicates = [k for k, v in dup.items() if v > 1]
            raise ValueError(
                "Les alias des étudiants doivent être uniques. "
                f"Vérifiez la liste des étudiants pour {duplicates}"
            )

        # Toutes les équipes doivent avoir un étudiant référent unique
        teams = {student.team for student in self.students if student.team}
        teams_ref = [student.team for student in self.students
                     if student.team and student.is_team_reference]
        if sorted(teams_ref) != sorted(list(teams)):
            # Find the team references that are not unique
            d = {}
            for x in teams_ref:
                if x not in d:
                    d[x] = 0
                d[x] += 1
            duplicates = [k for k, v in d.items() if v > 1]
            # Find the teams that have no team reference
            no_ref_teams = [team for team in teams if team not in teams_ref]
            bad = duplicates + no_ref_teams
            raise ValueError(
                "Chaque équipe doit avoir un étudiant référent unique. "
                f"Vérifiez la liste des étudiants pour {bad}"
            )
        return self

    def find_student(self, omnivox_code: str) -> Student | None:
        """
        Trouve un étudiant dans la liste à partir de son code omnivox.
        """
        for student in self.students:
            if student.omnivox_code == omnivox_code:
                return student
        return None

    def find_other_team_members(self, student: Student) -> list[Student]:
        """
        Trouve les autres membres de l'équipe d'un étudiant.

        Retourne une liste d'étudiants qui sont dans la même équipe que l'étudiant
        donné, à l'exception de l'étudiant lui-même.
        """
        return [
            s for s in self.students
            if s.team == student.team and s.omnivox_code != student.omnivox_code
        ]

    def __iter__(self): # type: ignore
        """
        Permet d'itérer sur les étudiants.
        """
        return iter(self.students)

    @classmethod
    def from_file(cls, path: str | Path) -> Self:
        """
        Lit les étudiants à partir d'un fichier CSV ou Excel et retourne une instance de Students.
        """
        return cls(students=read_student(path))

    def copy(self) -> "Students": # type: ignore
        """
        Retourne une copie de l'instance.
        """
        return Students(students=[student.copy() for student in self.students])

def read_student(path: str | Path) -> list[Student]:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return read_student_csv(path)
    elif path.suffix.lower() in {".xlsx", ".xls"}:
        return read_student_excel(path)
    else:
        raise ValueError(
            f"Le format de fichier {path.suffix} n'est pas supporté. "
            "Utilisez un fichier CSV ou Excel."
        )

def read_student_csv(path: str | Path) -> list[Student]:
    """
    Lit un fichier CSV contenant des informations sur les étudiants et retourne
    une liste d'instances de Student.
    """
    students = []
    with open(path, encoding="utf-8") as file:
        csv_data = csv.DictReader(file)
        for row in csv_data:
            students.append(Student.from_dict(row))
    return students

def read_student_excel(path: str | Path) -> list[Student]:
    """
    Lit un fichier Excel contenant des informations sur les étudiants et retourne
    une liste de dictionnaires représentant chaque étudiant.
    La structure attendue est la même que pour le fichier CSV.
    """
    # Nous évitons d'utiliser pandas pour minimiser les dépendances et parce que,
    # parfois au Collège de Maisonneuve, il n'est pas installé.
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True)) # type: ignore
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    students = []
    for row in rows[1:]:
        student = {headers[i]: (str(cell) if cell is not None else "")
                   for i, cell in enumerate(row)}
        if is_empty_row(student):
            continue
        students.append(Student.from_dict(student))
    return students

def is_empty_row(row: dict[str, str]) -> bool:
    """
    Vérifie si une ligne de données d'étudiant est vide.
    """
    return all(value.strip() == "" for value in row.values())
