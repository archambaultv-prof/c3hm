from pathlib import Path

import openpyxl as pyxl
import openpyxl.utils as pyxl_utils
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

from c3hm.core.generate.generate_word import scale_color_schemes
from c3hm.core.rubric import Rubric


def generate_excel_from_rubric(rubric: Rubric, output_path: Path | str) -> None:
    """
    Génère un document Excel servant à la correction à partir d’une Rubric
    """
    wb = pyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False

    # Tailles de colonnes
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    for i in range(3, 11):
        ws.column_dimensions[pyxl_utils.get_column_letter(i)].width = 15
    ws.column_dimensions[pyxl_utils.get_column_letter(11)].width = 60

    # Titre
    ws.append([rubric.title()])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.style = "Title"
    ws.append([])

    # Total sur X pts
    pts = "pt" if rubric.grid.total_score == 1 else "pts"
    ws.append([f"Total sur {rubric.grid.total_score} {pts}"])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.style = "Headline 2"

    total_row = 3
    total_cell = 2
    cell = ws.cell(row=ws.max_row, column=2)
    cell.style = "Calculation"

    # Grille
    grid = rubric.grid
    scale_colors = scale_color_schemes(len(grid.scale))

    # En-tête - seuils
    col_before_scale = 2
    ws.append([''] * col_before_scale + grid.thresholds)
    threshold_perfect_cell = "$C$4"
    threshold_row = ws.max_row
    for i in range(col_before_scale + 1, len(grid.thresholds) + col_before_scale + 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.style = "Explanatory Text"

    # En-tête - barème
    extra_cols = ["note calculée", "note manuelle", "note en pts", "commentaires"]
    ws.append([''] * col_before_scale + grid.scale + extra_cols)
    for i in range(col_before_scale + 1, len(grid.scale) + col_before_scale + 1 + len(extra_cols)):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.style = "Headline 4"
        if scale_colors and i - col_before_scale - 1 < len(scale_colors):
            color = scale_colors[i - col_before_scale - 1]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


    pts_letter = "B"
    criterion_rows = []
    computed_col = col_before_scale + len(grid.scale) + 1
    computed_letter = pyxl_utils.get_column_letter(computed_col)
    manual_col = col_before_scale + len(grid.scale) + 2
    manual_letter = pyxl_utils.get_column_letter(manual_col)
    final_col = col_before_scale + len(grid.scale) + 3
    final_letter = pyxl_utils.get_column_letter(final_col)
    for criterion in grid.criteria:
        # Critère
        ws.append([criterion.name, criterion.weight])
        cr = ws.max_row
        criterion_rows.append(cr)
        cell = ws.cell(row=ws.max_row, column=1)
        cell.style = "Headline 3"

        # note calculée
        cell = ws.cell(row=ws.max_row, column= computed_col)
        cell.value = (f"=sum({computed_letter}{cr+1}:"
                      f"{computed_letter}{cr+len(criterion.indicators)})")
        cell.style = "Calculation"

        # note manuelle
        cell = ws.cell(row=ws.max_row, column= manual_col)
        cell.style = "Input"

        # note en pts
        cell = ws.cell(row=ws.max_row, column= col_before_scale + len(grid.scale) + 3)
        computed_or_manual = (f"IF(ISBLANK({manual_letter}{cr}),{computed_letter}{cr},"
                              f"{manual_letter}{cr})")
        cell.value = f"={computed_or_manual}/{threshold_perfect_cell}*{pts_letter}{cr}"
        cell.style = "Calculation"

        # Descripteur
        for indicator in criterion.indicators:
            ws.append([indicator.name, 1])
            for i in range(2):
                cell = ws.cell(row=ws.max_row, column=i+1)
                cell.style = "Explanatory Text"
            # Affichage conditionnel pour les descripteurs
            for i in range(len(indicator.descriptors)):
                color = "D3D3D3"  # Gris clair
                if scale_colors:
                    color = scale_colors[i]
                col_letter = pyxl_utils.get_column_letter(i + col_before_scale + 1)
                rule = FormulaRule(
                    formula=[f'NOT(ISBLANK({col_letter}{ws.max_row}))'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
                )
                ws.conditional_formatting.add(
                    f'{col_letter}{ws.max_row}',
                    rule
                )

            # Note pour l'indicateur.
            # Si on utilise une formule trop avancée, Excel
            # va ajouter un @ sur certains ranges, ce qui va casser la formule.
            # Donc on utilise des formules simples sans confusion avec les
            # dynamic arrays.
            r = ws.max_row
            def cell_addr(i, r=r):
                return f"{pyxl_utils.get_column_letter(i + col_before_scale + 1)}{r}"
            def thr_addr(i):
                return f"{pyxl_utils.get_column_letter(i + col_before_scale + 1)}{threshold_row}"
            one_cell = [f"IF(ISBLANK({cell_addr(i)}),0,1)"
                        for i in range(len(indicator.descriptors))]
            one_cell = f"{" + ".join(one_cell)} = 1"
            val = [f"IF(ISBLANK({cell_addr(i)}),0,"
                   f"IF(ISNUMBER({cell_addr(i)}),{cell_addr(i)},{thr_addr(i)}))"
                   for i in range(len(indicator.descriptors))]
            val = f"{" + ".join(val)}"
            cell = ws.cell(row=ws.max_row, column=computed_col)
            cell.value = (f"=IF({one_cell},{val},NA())")
            cell.style = "Output"


    # Formule pour le total
    f = "=sum("
    f += ",".join([f"{final_letter}{r}" for r in criterion_rows])
    f += ")"
    ws.cell(row=total_row, column=total_cell).value = f

    wb.save(output_path)
