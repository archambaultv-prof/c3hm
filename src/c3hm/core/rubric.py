from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl as pyxl
from openpyxl import Workbook
from pydantic import BaseModel, Field

from c3hm.core.utils import has_max_decimals, split_decimal


class Indicator(BaseModel):
    """
    Représente un indicateur d'évaluation pour un critère donné.
    """
    name: str = Field(..., min_length=1)
    descriptors: list[str]


class Criterion(BaseModel):
    """
    Représente un critère d'évaluation. Ce dernière peut être formée d'une liste d'indicateurs.
    """
    name: str = Field(..., min_length=1)
    indicators: list[Indicator]
    weight: Decimal | None = Field(
        default=None,
        ge=Decimal(0)
    )

    def nb_indicators(self) -> int:
        """
        Retourne le nombre d'indicateurs dans le critère.
        """
        return len(self.indicators)


class RubricGrid(BaseModel):
    """
    Représente la grille d'évaluation.

    Une grille est formée de niveaux (Excellent, Très bien, ...) et d'une liste de critères.
    Il est possible d'ajouter des seuils pour chaque critère.
    """
    scale: list[str] = Field(
        default=["Excellent", "Très bien", "Bien", "Passable", "Insuffisant"],
        min_length=2
        )
    criteria: list[Criterion] = Field(..., min_length=1)
    total_score: Decimal = Field(
        default=100,
        ge=0
    )
    pts_precision: int = Field(
        default=0,
        ge=0
    )
    thresholds: list[int] = Field(
        default=[100, 85, 70, 60, 0],
        min_length=2,
    ),
    thresholds_precision: int = Field(
        default=0,
        ge=0
    )
    def nb_criteria(self) -> int:
        """
        Retourne le nombre de critères dans la grille.
        """
        return len(self.criteria)

    def nb_rows(self) -> int:
        """
        Retourne le nombre total de lignes dans la grille.

        Cela inclut une ligne pour le barème, une ligne pour chaque critère
        et une ligne pour chaque indicateur de chaque critère.
        """
        return 1 + self.nb_criteria() + sum(
            criterion.nb_indicators() for criterion in self.criteria
        )

    def nb_columns(self) -> int:
        """
        Retourne le nombre total de colonnes dans la grille.

        Cela inclut une colonne pour les critères/descripteurs et une colonne
        pour chaque niveau de barème.
        """
        return 1 + len(self.scale)

    def validate(self) -> None:
        """
        Valide la grille d'évaluation.

        Vérifie que le nombre de seuils correspond au nombre de niveaux du barème,
        que les poids des critères sont valides et que les descripteurs et
        indicateurs sont correctement définis.
        """
        # Vérification du nombre de seuils
        if len(self.scale) != len(self.thresholds):
            raise ValueError("Le nombre de seuils doit correspondre "
                             "au nombre de niveaux du barème.")

        # Vérification des poids
        weight_total = Decimal(0)
        nb_without_weight = 0
        for criterion in self.criteria:
            if criterion.weight is None:
                nb_without_weight += 1
            else:
                # Verification si precision est respectée

                if not has_max_decimals(criterion.weight, self.pts_precision):
                    raise ValueError(
                        f"Le poids du critère '{criterion.name}' "
                        f"doit avoir au plus {self.pts_precision} décimales."
                    )
                weight_total += criterion.weight
        unallocated_weight = self.total_score - weight_total
        if unallocated_weight < 0:
            raise ValueError("La somme des poids des critères dépasse le total.")
        if nb_without_weight > 0:
            weights = split_decimal(unallocated_weight, nb_without_weight, self.pts_precision)
            weights.reverse()
            for criterion in self.criteria:
                if criterion.weight is None:
                    criterion.weight = weights.pop()

        # Vérification des descripteurs et indicateurs
        for criterion in self.criteria:
            if not criterion.indicators:
                raise ValueError(f"Le critère '{criterion.name}' n'a pas d'indicateurs.")
            for indicator in criterion.indicators:
                if len(indicator.descriptors) != len(self.scale):
                    raise ValueError(
                        f"Le nombre de descripteurs pour l'indicateur '{indicator.name}' "
                        f"doit correspondre au nombre de niveaux du barème."
                    )


class Student(BaseModel):
    """
    Représente un étudiant avec un nom, prénom et code Omnivox.
    """
    first_name: str = Field(..., min_length=1)
    last_name: str = Field(..., min_length=1)
    omnivox_code: str = Field(..., min_length=1)
    id: str = Field(..., min_length=1)

    def ws_name(self) -> str:
        """
        Retourne le nom de la feuille de calcul pour l'étudiant.

        Le nom est formé du prénom et du nom, séparés par un espace.
        """
        return self.id

class Rubric(BaseModel):
    """
    Représente une grille d'évaluation pour un cours et une évaluation spécifiques.
    """
    course: str | None = Field(
         default=None,
         min_length=1
         )
    evaluation: str | None  = Field(
        default=None,
        min_length=1
        )
    grid: RubricGrid
    students: list[Student] = Field(
        default_factory=list,
        min_length=0
    )

    def title(self) -> str:
        """
        Retourne le titre de la grille d'évaluation.

        Le titre est formé du nom du cours et de l'évaluation, séparés par un tiret.
        """
        endash = "\u2013"
        title = "Grille d'évaluation"
        if self.evaluation:
            title += f" {endash} {self.evaluation}"
        if self.course:
            title += f" {endash} {self.course}"
        return title

    def validate_rubric(self) -> None:
        """
        Valide la grille d'évaluation et les étudiants.

        Voir la méthode validate de RubricGrid pour les détails de la validation.
        """
        self.grid.validate()

        if len(self.students) != len(set(s.id for s in self.students)):
            raise ValueError("Les étudiants doivent avoir des identifiants uniques.")


def load_rubric_from_xlsx(
        filepath: str | Path,
         *,
         sheet_name="c3hm",
         ) -> Rubric:
    """
    Charge et initialise une grille d'évaluation à partir d'un fichier Excel.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Le fichier '{filepath}' n'existe pas.")

    wb = pyxl.load_workbook(filepath, data_only=True)
    config = read_c3hm_config_xl(wb, sheet_name=sheet_name)
    students = read_c3hm_students_xl(wb, sheet_name=sheet_name)

    try:
        rubric = rubric_from_config(config, students)
    except Exception as e:
        raise ValueError(f"Erreur lors du chargement du fichier : {filepath}") from e

    return rubric

def find_named_cell(wb: Workbook,
                    sheet_name: str,
                    named_cell: str) -> tuple[int, int]:
    """
    Trouve la position d'une cellule nommée dans une feuille de calcul.
    Retourne la ligne et la colonne de la cellule.
    """
    # Vérifie si la feuille existe
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' non trouvée dans le classeur.")
    ws = wb[sheet_name]

    # Vérifie si la cellule nommée existe
    defn = None
    if named_cell in wb.defined_names:
        defn = wb.defined_names[named_cell]
    elif named_cell in ws.defined_names:
        defn = ws.defined_names[named_cell]

    if defn is None:
        raise ValueError(f"Cellule nommée '{named_cell}' non trouvée dans le classeur.")

    for title, coord in defn.destinations:
        if title == sheet_name:
            cell = ws[coord]
            return cell.row, cell.column
    raise ValueError(f"Cellule nommée '{named_cell}' non trouvée dans la feuille '{sheet_name}'.")


def read_c3hm_students_xl(
        wb: Workbook,
        *,
        sheet_name="c3hm",
        named_cell="cthm_code_omnivox"
        ) -> list[dict]:
    """
    Lit la liste des étudiants à partir d'un fichier Excel.
    Part depuis une cellule nommée et lit vers le bas jusqu'à la première
    cellule vide. Assume que le code est dans la première colonne, le
    prénom dans la deuxième et le nom dans la troisième.

    Retourne une liste de dictionnaires.
    """
    students = []

    cell_row, cell_col = find_named_cell(wb, sheet_name, named_cell)

    # On commence à lire à la ligne suivante
    # et on lit jusqu'à la première cellule vide
    row = cell_row + 1
    ws = wb[sheet_name]
    while True:
        code_cell = ws.cell(row=row, column=cell_col)
        first_name_cell = ws.cell(row=row, column=cell_col + 1)
        last_name_cell = ws.cell(row=row, column=cell_col + 2)
        id_cell = ws.cell(row=row, column=cell_col + 3)

        if code_cell.value is None:
            break

        d = {
            "omnivox_code": str(code_cell.value),
            "first_name": str(first_name_cell.value),
            "last_name": str(last_name_cell.value),
            "id": str(id_cell.value)
        }
        students.append(d)
        row += 1

    return students

def read_c3hm_config_xl(
        wb: Workbook,
        *,
        sheet_name="c3hm",
        named_cell="cthm_config_key"
        ) -> list[tuple[str, Any]]:
    """
    Lit la configuration d'une grille d'évaluation à partir d'un fichier Excel.
    Part depuis une cellule nommée et lit vers le bas jusqu'à la première
    cellule vide. Assume que la cellule nommée est dans la première colonne et
    que les valeurs sont dans la deuxième colonne.
    """
    config = []

    cell_row, cell_col = find_named_cell(wb, sheet_name, named_cell)

    # On commence à lire à la ligne suivante
    # et on lit jusqu'à la première cellule vide
    row = cell_row + 1
    ws = wb[sheet_name]
    while True:
        key_cell = ws.cell(row=row, column=cell_col)
        value_cell = ws.cell(row=row, column=cell_col + 1)

        if key_cell.value is None:
            break

        key = str(key_cell.value)
        value = value_cell.value
        config.append((key, value))
        row += 1

    return config


def rubric_from_config(config: list[tuple[str, Any]], students: list[dict]) -> Rubric:
    """
    Charge et initialise une grille d'évaluation à partir d'un configuration.

    Voir le gabarit grille_5_niveaux.xlsx pour la structure attendue.
    """

    check_duplicate_param(config)

    cours = get_param(config, "Cours")
    evaluation = get_param(config, "Évaluation")
    total = get_param(config, "Total", 100)
    pts_precision = get_param(config, "Précision poids", 0)
    scale = get_scale(config)
    thresholds = get_thresholds(config)
    thresholds_precision = get_param(config, "Précision seuils", 0)
    criteria = get_criteria(config)
    if not criteria:
        raise ValueError("Aucun critère trouvé.")
    students = [Student(**d) for d in students]

    grid = RubricGrid(
        scale=scale,
        criteria=criteria,
        total_score=total,
        pts_precision=pts_precision,
        thresholds=thresholds,
        thresholds_precision=thresholds_precision
    )

    rubric = Rubric(
        course=cours,
        evaluation=evaluation,
        grid=grid,
        students=students
    )

    rubric.validate_rubric()
    return rubric

def check_duplicate_param(config: list[tuple[str, Any]]):
    """
    Vérifie si un paramètre est défini plusieurs fois dans la configuration.

    Les paramètres Critère, Indicateur et Descripteur sont autorisés à être
    définis plusieurs fois.
    """
    seen = set()
    ok_duplicate = ["Critère", "Indicateur", "Descripteur"]
    for k, _ in config:
        skip = False
        for x in ok_duplicate:
            if k.startswith(x):
                skip = True
                break
        if skip:
            continue

        if k in seen:
            raise ValueError(f"Le paramètre '{k}' est défini plusieurs fois.")
        seen.add(k)


def get_param(c: list[tuple[str, Any]], param: str, default: Any = None) -> Any:
    """
    Retourne la valeur d'un paramètre à partir d'une liste de tuples.
    """
    for k, v in c:
        if k == param:
            return v
    return default


def get_scale(c: list[tuple[str, Any]]) -> list[str]:
    """
    Retourne la liste des niveaux à partir de la configuration
    ou la liste par défaut si aucun niveau n'est trouvé.
    """
    scale = []
    current_level = 0
    for k, v in c:
        if k.startswith("Niveau "):
            level = int(k[7:])
            if level != current_level + 1:
                raise ValueError(f"Niveau '{level}' non consécutif.")
            current_level = level
            scale.append(v)
    if not scale:
        return ["Excellent", "Très bien", "Bien", "Passable", "Insuffisant"]
    else:
        return scale


def get_thresholds(c: list[tuple[str, Any]]) -> list[int]:
    """
    Retourne la liste des seuils à partir de la configuration
    ou la liste par défaut si aucun seuil n'est trouvé.
    """
    thresholds = []
    current_level = 0
    for k, v in c:
        if k.startswith("Seuil "):
            level = int(k[6:])
            if level != current_level + 1:
                raise ValueError(f"Seuil '{level}' non consécutif.")
            current_level = level
            thresholds.append(v)
    if not thresholds:
        return [100, 85, 70, 60, 0]
    else:
        return thresholds

def get_criteria(c: list[tuple[str, Any]]) -> list[Criterion]:
    """
    Retourne la liste des critères à partir d'une configuration.

    Lit la configuration dans l'ordre.
    """
    criteria: list[Criterion] = []
    current_criterion: Criterion | None = None
    current_ind: Indicator | None = None
    current_desc_number: int = 0
    seen_weight = False

    for key, value in c:
        if key == "Critère":
            if current_criterion:
                if current_ind is not None:
                    current_criterion.indicators.append(current_ind)
                    current_ind = None

                criteria.append(current_criterion)
            # Nouveau critère
            current_criterion = Criterion(name=value, indicators=[])
            seen_weight = False

        elif key == 'Critère poids':
            if current_criterion:
                if seen_weight:
                    raise ValueError(
                        f"Critère poids défini plusieurs fois pour {current_criterion.name}."
                    )
                if value:
                    current_criterion.weight = Decimal(str(value))
                seen_weight = True
            else:
                raise ValueError("Critère poids sans critère.")

        elif key == "Indicateur":
            if current_criterion is None:
                raise ValueError("Indicateur sans critère.")
            if current_ind is not None:
                current_criterion.indicators.append(current_ind)
            current_ind = Indicator(name=value, descriptors=[])
            current_desc_number = 0

        elif key.startswith("Descripteur "):
            if current_ind is None:
                raise ValueError("Descripteur sans indicateur.")
            n = int(key[12:])
            if n != current_desc_number + 1:
                raise ValueError(f"Descripteur '{key}' non consécutif.")
            current_desc_number = n
            current_ind.descriptors.append(value)

    if current_criterion:
        if current_ind is not None:
            current_criterion.indicators.append(current_ind)
        criteria.append(current_criterion)

    return criteria

